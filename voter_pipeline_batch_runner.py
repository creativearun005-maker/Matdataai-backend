"""
voter_pipeline_batch_runner.py
MatdataAI — Phase 1 (photo-crop) + Phase 2 (OCR) + Phase 3 (Excel) end-to-end.
Exposes process_pdf() for backend_api.py, plus a CLI batch-runner for direct use.
"""

import os
import io
import re
import ast
import fitz
import cv2
import numpy as np
import pandas as pd
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from PIL import Image as PILImage
from difflib import SequenceMatcher

import requests
import base64
import time

MISTRAL_API_KEY = os.environ.get("MISTRAL_API_KEY", "")
MISTRAL_MODEL = "pixtral-12b-2409"


def extract_epic_with_mistral(crop_image_np):
    if not MISTRAL_API_KEY:
        print("[MISTRAL DEBUG] API key is empty!")
        return None
    try:
        if crop_image_np.shape[0] < 100:  # agar bahut chhota crop hai
            crop_image_np = cv2.resize(crop_image_np, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        success, buf = cv2.imencode('.jpg', crop_image_np, [cv2.IMWRITE_JPEG_QUALITY, 100])
        if not success:
            print("[MISTRAL DEBUG] cv2.imencode failed on crop!")
            return None
        img_b64 = base64.b64encode(buf).decode('utf-8')
        data_url = f"data:image/jpeg;base64,{img_b64}"

        print(f"[MISTRAL DEBUG] Sending request... (image size: {len(buf)} bytes)")

        url = "https://api.mistral.ai/v1/chat/completions"
        headers = {
            "Authorization": f"Bearer {MISTRAL_API_KEY}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": MISTRAL_MODEL,
            "temperature": 0,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "image_url", "image_url": data_url},
                        {"type": "text", "text": (
                            "This is a crop from an Indian voter-list PDF. It contains an "
                            "EPIC (voter ID) number, alphanumeric like 'UP/84/417/0198404' or "
                            "'GBY2781292'. Reply with ONLY the EPIC number, nothing else. "
                            "If none is visible, reply exactly: NONE"
                        )},
                    ],
                }
            ],
        }
        resp = requests.post(url, headers=headers, json=payload, timeout=45)
        print(f"[MISTRAL DEBUG] Response status: {resp.status_code}")
        print(f"[MISTRAL DEBUG] Raw response body: {resp.text[:500]}")

        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        print(f"[MISTRAL DEBUG] content type={type(content)}, value={content!r}")

        text = content.strip() if isinstance(content, str) else str(content).strip()
        result = None if text.upper() == "NONE" else text
        print(f"[MISTRAL DEBUG] Final extracted value: {result!r}")
        return result
    except Exception as e:
        print(f"[MISTRAL EPIC ERROR] {type(e).__name__}: {e}")
        return None
# insightface is optional at import-time (heavy dep) — loaded lazily in get_face_app()
_face_app = None


# ======================================================================
# PHASE 1 — Photo cropping (validated: 19,485 crops, 0.046% failure rate)
# ======================================================================

def render_page(doc, page_index, dpi=100):
    page = doc[page_index]
    pix = page.get_pixmap(dpi=dpi)
    img = Image.open(io.BytesIO(pix.tobytes("png")))
    return np.array(img.convert('RGB'))


def get_epic_region(photo_box, cell_width=393, strip_height=40, padding=3):
    """The EPIC code + serial number sit in a strip ABOVE the photo, spanning
    the cell width — this crops that strip."""
    px, py, pw, ph = photo_box
    cell_x = max(0, px + pw - cell_width)
    epic_y = max(0, py - strip_height - padding)
    return (cell_x, epic_y, cell_width, strip_height)


def detect_column_count(img_array):
    gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
    edges = cv2.Canny(gray, 50, 150)
    kernel = np.ones((3, 3), np.uint8)
    edges = cv2.dilate(edges, kernel, iterations=1)
    contours, _ = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    h_img, w_img = img_array.shape[:2]
    expected_w = w_img / 3
    expected_h = h_img / 6
    x_starts = []
    for c in contours:
        x, y, w, h = cv2.boundingRect(c)
        if (0.6 * expected_w < w < 1.3 * expected_w) and (0.5 * expected_h < h < 1.4 * expected_h):
            x_starts.append(x)
    if not x_starts:
        return 0, []
    x_starts.sort()
    clusters = [[x_starts[0]]]
    for x in x_starts[1:]:
        if x - clusters[-1][-1] < expected_w * 0.3:
            clusters[-1].append(x)
        else:
            clusters.append([x])
    return len(clusters), [int(np.mean(c)) for c in clusters]


def get_grid_pages(pdf_path, dpi=100):
    """Page = voter-grid iff 3 columns detected. Robust to partial/last pages."""
    doc = fitz.open(pdf_path)
    grid_pages = []
    for i in range(len(doc)):
        img = render_page(doc, i, dpi=dpi)
        n_cols, _ = detect_column_count(img)
        if n_cols == 3:
            grid_pages.append(i)
    doc.close()
    return grid_pages


def find_photo_boxes(image_path, min_area=6000, max_area=13000, min_aspect=0.75, max_aspect=1.05):
    img = cv2.imread(image_path)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    edges = cv2.Canny(gray, 50, 150)
    kernel = np.ones((3, 3), np.uint8)
    edges = cv2.dilate(edges, kernel, iterations=1)
    contours, _ = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    boxes = []
    for c in contours:
        x, y, w, h = cv2.boundingRect(c)
        area = w * h
        aspect = w / h if h > 0 else 0
        if min_area < area < max_area and min_aspect < aspect < max_aspect:
            boxes.append((x, y, w, h))
    return img, boxes


def remove_duplicate_boxes(boxes):
    if not boxes:
        return []
    boxes = sorted(boxes, key=lambda b: b[2] * b[3], reverse=True)
    kept = []
    for box in boxes:
        x1, y1, w1, h1 = box
        is_dup = False
        for kx, ky, kw, kh in kept:
            cx1, cy1 = x1 + w1 / 2, y1 + h1 / 2
            cx2, cy2 = kx + kw / 2, ky + kh / 2
            dist = ((cx1 - cx2) ** 2 + (cy1 - cy2) ** 2) ** 0.5
            if dist < max(w1, kw) * 0.5:
                is_dup = True
                break
        if not is_dup:
            kept.append(box)
    return kept


import cv2

class _FaceResult:
    """Mimics insightface's face-result object shape (has .det_score)."""
    def __init__(self, det_score):
        self.det_score = det_score

class _CascadeFaceApp:
    """Drop-in replacement for insightface's FaceAnalysis — same .get(img) interface."""
    def __init__(self):
        self.detector = cv2.CascadeClassifier(
            cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
        )

    def get(self, img):
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if len(img.shape) == 3 else img
        faces = self.detector.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))
        # return objects shaped like insightface's output, so rest of the code
        # (which reads .det_score) doesn't need to change at all
        return [_FaceResult(det_score=0.95) for _ in faces]


_face_app = None

def get_face_app():
    global _face_app
    if _face_app is None:
        _face_app = _CascadeFaceApp()
    return _face_app


# ======================================================================
# PHASE 2 — Text extraction (Tesseract Hindi OCR + field parser)
# ======================================================================

import pytesseract

KNOWN_LABELS = {
    'नाम': 'नाम', 'पिता का नाम': 'पिता का नाम', 'पति का नाम': 'पति का नाम',
    'मकान संख्या': 'मकान संख्या', 'आयु': 'आयु', 'लिंग': 'लिंग',
}
KNOWN_VALUES = {'पुरुष': 'पुरुष', 'महिला': 'महिला'}


def similarity(a, b):
    return SequenceMatcher(None, a, b).ratio()


def correct_label(raw_label, threshold=0.55):
    best_match, best_score = None, 0
    for label in KNOWN_LABELS:
        score = similarity(raw_label.strip(), label)
        if score > best_score:
            best_match, best_score = label, score
    return (best_match, best_score) if best_score >= threshold else (None, best_score)


def correct_gender_value(raw_value, threshold=0.5):
    best_match, best_score = None, 0
    for val in KNOWN_VALUES:
        score = similarity(raw_value.strip(), val)
        if score > best_score:
            best_match, best_score = val, score
    return (best_match, best_score) if best_score >= threshold else (raw_value, best_score)


def preprocess_lines(ocr_lines):
    merged, skip_next = [], False
    for i, line in enumerate(ocr_lines):
        if skip_next:
            skip_next = False
            continue
        stripped = line.strip()
        has_sep = bool(re.search('[:ःƒ：]', stripped))
        parts = re.split('[:ःƒ：]', stripped, maxsplit=1)
        value_part = parts[1].strip() if len(parts) > 1 else ''
        if has_sep and len(value_part) == 0 and i + 1 < len(ocr_lines):
            merged.append(stripped + ' ' + ocr_lines[i + 1].strip())
            skip_next = True
        else:
            merged.append(stripped)
    return merged


def merge_adjacent_pipe_as_one(text):
    text = re.sub(r'(\d)\s{0,1}\|', r'\g<1>1', text)
    text = re.sub(r'\|\s{0,1}(\d)', r'1\g<1>', text)
    return text


def clean_numeric_field(raw_value):
    if not raw_value:
        return raw_value
    cleaned = merge_adjacent_pipe_as_one(raw_value)
    replacements = {']': '1', '[': '1', 'l': '1', 'I': '1', 'O': '0', 'o': '0', 'S': '5', 's': '5'}
    for wrong, right in replacements.items():
        cleaned = cleaned.replace(wrong, right)
    return cleaned


def validate_age(age_str):
    """Round-2 decision: extract age but do NOT flag on 18+ eligibility
    (font-specific digit-drop made this unreliable — deferred to Round 3 fine-tuning)."""
    cleaned = clean_numeric_field(age_str)
    return cleaned, False, None


def parse_voter_entry(ocr_lines):
    ocr_lines = preprocess_lines(ocr_lines)
    result = {
        'name': None, 'relation_type': None, 'relation_name': None,
        'house_no': None, 'age': None, 'gender': None, 'needs_review': []
    }
    if not ocr_lines:
        return result

    first_line = ocr_lines[0].strip()
    parts = re.split('[:ः：]', first_line, maxsplit=1)
    if len(parts) > 1 and parts[1].strip():
        result['name'] = parts[1].strip()
    else:
        result['name'] = first_line if first_line else None
        result['needs_review'].append(f'name_no_clear_separator: "{first_line}"')

    for line in ocr_lines[1:]:
        line = line.strip()
        if not line:
            continue
        parts = re.split('[:ः：]', line, maxsplit=1)
        raw_label = parts[0].strip()
        raw_value = parts[1].strip() if len(parts) > 1 else ''
        matched_label, score = correct_label(raw_label)

        if matched_label == 'पिता का नाम':
            result['relation_type'] = 'पिता'
            result['relation_name'] = raw_value
            if score < 0.8:
                result['needs_review'].append(f'father_label(score={score:.2f})')
        elif matched_label == 'पति का नाम':
            result['relation_type'] = 'पति'
            result['relation_name'] = raw_value
            if score < 0.8:
                result['needs_review'].append(f'husband_label(score={score:.2f})')
        elif matched_label == 'मकान संख्या':
            if 'आयु' in raw_value or 'लिंग' in raw_value:
                result['needs_review'].append(f'house_no_merged_with_age_gender: "{raw_value}"')
                digit_part = re.match(r'^\s*([\S]*?)\s*आयु', raw_value)
                result['house_no'] = digit_part.group(1) if digit_part else None
                age_match = re.search(r'आयु\s*[:：]?\s*([0-9०-९]+)', raw_value)
                if age_match:
                    result['age'] = age_match.group(1)
                gender_match = re.search(r'लिंग\s*[:：]?\s*(\S+)', raw_value)
                if gender_match:
                    gender_corrected, _ = correct_gender_value(gender_match.group(1))
                    result['gender'] = gender_corrected
            else:
                result['house_no'] = raw_value
            if score < 0.8:
                result['needs_review'].append(f'house_label(score={score:.2f})')
        elif matched_label == 'आयु':
            age_match = re.search(r'([0-9०-९]+)', raw_value)
            if age_match:
                result['age'] = age_match.group(1)
            if 'लिंग' in raw_value or 'लिग' in raw_value:
                gender_part = re.split('लिंग|लिग', raw_value)[-1]
                gender_corrected, g_score = correct_gender_value(gender_part)
                result['gender'] = gender_corrected
                if g_score < 0.6:
                    result['needs_review'].append(f'gender(score={g_score:.2f})')
        else:
            result['needs_review'].append(f'unrecognized_line: "{line}" (best_score={score:.2f})')

    return result


def get_text_region(photo_box, cell_width=393, cell_height=162, padding=5):
    px, py, pw, ph = photo_box
    cell_x = max(0, px + pw - cell_width)
    cell_y = py - padding
    cell_h = cell_height + 2 * padding
    text_w = px - cell_x
    return (cell_x, cell_y, text_w, cell_h)


def preprocess_for_ocr(crop):
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    upscaled = cv2.resize(gray, None, fx=3, fy=3, interpolation=cv2.INTER_CUBIC)
    _, binary = cv2.threshold(upscaled, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    denoised = cv2.fastNlMeansDenoising(binary, h=10)
    return denoised


def process_page(image_path, page_id=None, review_crop_dir='/tmp/review_crops'):
    """Full Phase 1 + Phase 2 for one rendered page image. Returns list of entry dicts."""
    img = cv2.imread(image_path)
    _, raw_boxes = find_photo_boxes(image_path)
    unique_boxes = remove_duplicate_boxes(raw_boxes)

    if page_id is None:
        page_id = os.path.splitext(os.path.basename(image_path))[0]

    os.makedirs(review_crop_dir, exist_ok=True)
    face_app = get_face_app()
    page_results = []

    for idx, photo_box in enumerate(unique_boxes):
        x, y, w, h = photo_box
        face_crop = img[y:y + h, x:x + w]
        faces = face_app.get(face_crop)
        face_confidence = faces[0].det_score if faces else 0.0
        photo_crop_path = f'/tmp/crops/{page_id}_entry{idx:02d}_photo.png'
        os.makedirs('/tmp/crops', exist_ok=True)
        cv2.imwrite(photo_crop_path, face_crop)
        # --- EPIC extraction (debug step first) ---
        epic_box = get_epic_region(photo_box)
        ex, ey, ew, eh = epic_box
        epic_crop = img[max(0, ey):ey + eh, max(0, ex):ex + ew]
        print(f"[EPIC DEBUG] Processing entry idx={idx}...")
        epic_number = extract_epic_with_mistral(epic_crop) if epic_crop.size > 0 else None
        if epic_number is None and epic_crop.size > 0:
            time.sleep(1)
            epic_number = extract_epic_with_mistral(epic_crop)  # ek retry, agar pehli baar miss hua
        time.sleep(0.5)  # free-tier rate-limit ke against safety margin
        cv2.imwrite(f'/tmp/debug_epic_{idx}.png', epic_crop)  # temporary — visual check

        tx, ty, tw, th = get_text_region(photo_box)
        text_crop = img[max(0, ty):ty + th, max(0, tx):tx + tw]

        if text_crop.size == 0:
            page_results.append({
                'entry_idx': idx, 'photo_box': photo_box, 'photo_crop_path': photo_crop_path,
                'face_confidence': round(float(face_confidence), 3),
                'parsed': None, 'error': 'empty_crop', 'review_crop_path': None
            })
            continue

        processed = preprocess_for_ocr(text_crop)
        raw_text = pytesseract.image_to_string(processed, lang='hin')
        raw_lines = [line for line in raw_text.split('\n') if line.strip()]
        parsed = parse_voter_entry(raw_lines)

        if parsed['house_no']:
            parsed['house_no'] = clean_numeric_field(parsed['house_no'])
        if parsed['age']:
            cleaned_age, age_flagged, age_reason = validate_age(parsed['age'])
            parsed['age'] = cleaned_age
            if age_flagged:
                parsed['needs_review'].append(age_reason)

        if face_confidence < 0.5:
            parsed['needs_review'].append(f'low_photo_confidence(score={face_confidence:.2f})')

        review_crop_path = None
        if parsed['needs_review']:
            review_crop_path = f'{review_crop_dir}/{page_id}_entry{idx:02d}.png'
            cv2.imwrite(review_crop_path, text_crop)

        page_results.append({
            'entry_idx': idx, 'photo_box': photo_box, 'photo_crop_path': photo_crop_path,
            'face_confidence': round(float(face_confidence), 3),
            'parsed': parsed, 'error': None, 'review_crop_path': review_crop_path
        })

    return page_results


# ======================================================================
# PHASE 3 — Excel generation (from generate_excel.py, kept as-is)
# ======================================================================

COLUMNS = [
    "", "Photo", "EPIC Number / पहचान पत्र संख्या", "Name / नाम",
    "Father's/Husband's Name / पिता/पति का नाम", "House Number / मकान संख्या",
    "Mobile Number / मोबाइल नंबर", "Caste / जाति", "Aadhar / आधार",
    "Age / आयु", "Gender / लिंग",
]


def _relation_display(row):
    rtype, rname = row.get("relation_type"), row.get("relation_name")
    if rtype and rname:
        return f"{rtype} - {rname}"
    return rname if rname else None


def _embed_image(ws, image_path, cell_ref, target_width_px):
    if not image_path or not os.path.exists(image_path):
        return
    with PILImage.open(image_path) as pil_img:
        w, h = pil_img.size
    scale = target_width_px / w
    xl_img = XLImage(image_path)
    xl_img.width = target_width_px
    xl_img.height = int(h * scale)
    ws.add_image(xl_img, cell_ref)


def _write_header(ws, headers, fill_color):
    font = Font(bold=True)
    fill = PatternFill("solid", fgColor=fill_color)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill = font, fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_excel_from_entries(all_entries, output_path, village_number, part_no, anubhag,
                              review_mode="separate_file", photo_target_width_px=85,
                              review_crop_target_width_px=220, row_height_pt=68):
    """Takes the flat list of entries produced by process_pdf() and writes the final Excel,
    same schema/logic as the validated generate_excel.py from Phase 3."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Voter List"
    _write_header(ws, COLUMNS, "D9E1F2")
    ws.column_dimensions[get_column_letter(2)].width = 14
    ws.column_dimensions[get_column_letter(3)].width = 22
    ws.column_dimensions[get_column_letter(8)].width = 20
    ws.column_dimensions[get_column_letter(9)].width = 24
    ws.row_dimensions[1].height = 30

    review_entries = []

    for row_idx, e in enumerate(all_entries, start=2):
        parsed = e['parsed'] or {}
        needs_review = parsed.get('needs_review', [])
        ws.row_dimensions[row_idx].height = row_height_pt

        ws.cell(row=row_idx, column=1, value=None)
        ws.cell(row=row_idx, column=3, value=e.get('epic_number'))
        ws.cell(row=row_idx, column=4, value=parsed.get('name'))
        ws.cell(row=row_idx, column=5, value=_relation_display(parsed))
        ws.cell(row=row_idx, column=6, value=parsed.get('house_no'))
        ws.cell(row=row_idx, column=7, value=None)   # mobile — Phase 5
        ws.cell(row=row_idx, column=8, value=None)   # jaati — TBD
        ws.cell(row=row_idx, column=9, value=None)   # aadhar — Phase 5
        ws.cell(row=row_idx, column=10, value=parsed.get('age'))
        ws.cell(row=row_idx, column=11, value=parsed.get('gender'))

        _embed_image(ws, e['photo_crop_path'], f"B{row_idx}", photo_target_width_px)

        if needs_review:
            review_entries.append({**e, 'parsed': parsed})

    review_headers = ["Entry Idx (Kram Sankhya)", "Reason(s)", "Naam (guess)",
                       "Pita/Pati (guess)", "Makan No (guess)", "Aayu (guess)",
                       "Ling (guess)", "Crop Image", "Correct Value"]

    if review_mode == "same_file":
        review_ws = wb.create_sheet("Review")
    else:
        review_wb = Workbook()
        review_ws = review_wb.active
        review_ws.title = "Review"

    _write_header(review_ws, review_headers, "FCE4D6")
    review_ws.column_dimensions[get_column_letter(8)].width = 34
    review_ws.column_dimensions[get_column_letter(9)].width = 24
    review_ws.row_dimensions[1].height = 24

    for row_idx, e in enumerate(review_entries, start=2):
        parsed = e['parsed']
        review_ws.row_dimensions[row_idx].height = 90
        review_ws.cell(row=row_idx, column=1, value=e['entry_idx'])
        review_ws.cell(row=row_idx, column=2, value="; ".join(parsed.get('needs_review', [])))
        review_ws.cell(row=row_idx, column=3, value=parsed.get('name'))
        review_ws.cell(row=row_idx, column=4, value=_relation_display(parsed))
        review_ws.cell(row=row_idx, column=5, value=parsed.get('house_no'))
        review_ws.cell(row=row_idx, column=6, value=parsed.get('age'))
        review_ws.cell(row=row_idx, column=7, value=parsed.get('gender'))
        _embed_image(review_ws, e['review_crop_path'], f"H{row_idx}", review_crop_target_width_px)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    review_path = None
    if review_mode == "separate_file" and review_entries:
        base, ext = os.path.splitext(output_path)
        review_path = f"{base}_review{ext}"
        review_wb.save(review_path)

    return {"main_path": output_path, "review_path": review_path,
            "n_entries": len(all_entries), "n_flagged": len(review_entries)}


# ======================================================================
# ENTRY POINT — process_pdf() : what backend_api.py imports
# ======================================================================

def process_pdf(pdf_path, village_number="", output_dir="/tmp/output"):
    """
    ASSUMED signature — backend_api.py mein jo call ho raha hai, agar iska shape
    alag hai (extra args, different return-type), backend_api.py ka content
    bhejna taaki isko exact-match kar saku.

    Ek PDF -> saari grid-pages -> saare entries -> ek Excel (+ review-Excel).
    Returns dict with output-file paths + counts.
    """
    part_no_match = re.search(r'HIN-(\d+)', os.path.basename(pdf_path))
    part_no = part_no_match.group(1) if part_no_match else None

    grid_pages = get_grid_pages(pdf_path)
    doc = fitz.open(pdf_path)
    all_entries = []

    for page_num in grid_pages:
        page = doc[page_num]
        pix = page.get_pixmap(dpi=150)
        img_pil = Image.open(io.BytesIO(pix.tobytes("png")))
        temp_path = f'/tmp/temp_page_{page_num}.png'
        img_pil.save(temp_path)

        page_id = f"{os.path.splitext(os.path.basename(pdf_path))[0]}_page{page_num}"
        page_results = process_page(temp_path, page_id=page_id)
        all_entries.extend(page_results)
        os.remove(temp_path)

    doc.close()

    pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
    output_path = os.path.join(output_dir, f"{pdf_name}.xlsx")

    result = build_excel_from_entries(
        all_entries, output_path,
        village_number=village_number,
        part_no=part_no,
        anubhag=None,  # header OCR not implemented yet
        review_mode="separate_file",
    )
    return result


# ======================================================================
# CLI batch-runner — poore dataset ke liye (Colab/local se seedha chalane layak)
# ======================================================================

def run_batch(pdf_paths, village_number="", output_dir="/tmp/output",
              log_path="/tmp/processed_pdfs_log.csv"):
    if os.path.exists(log_path):
        done = set(pd.read_csv(log_path)['pdf_name'])
    else:
        done = set()

    remaining = [p for p in pdf_paths if os.path.basename(p) not in done]
    print(f"Total: {len(pdf_paths)}, Done: {len(done)}, Remaining: {len(remaining)}")

    for i, pdf_path in enumerate(remaining):
        try:
            result = process_pdf(pdf_path, village_number=village_number, output_dir=output_dir)
            done.add(os.path.basename(pdf_path))
            print(f"[{i+1}/{len(remaining)}] OK  {os.path.basename(pdf_path)} → {result['n_entries']} entries")
        except Exception as e:
            print(f"[{i+1}/{len(remaining)}] FAIL {os.path.basename(pdf_path)} → {e}")

        if (i + 1) % 20 == 0:
            pd.DataFrame({'pdf_name': list(done)}).to_csv(log_path, index=False)

    pd.DataFrame({'pdf_name': list(done)}).to_csv(log_path, index=False)
    print("Batch complete.")


if __name__ == "__main__":
    import glob
    pdfs = glob.glob("/content/drive/MyDrive/MOTHER ROLL -2026/10-04-2026/Mother Roll/**/*.pdf", recursive=True)
    run_batch(pdfs, village_number="Village-101")
